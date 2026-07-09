import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

from config import AppConfig
from models import Transacao, TransacaoHotel

_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

MESES_PT = {
    "Jan": "01", "Fev": "02", "Mar": "03", "Abr": "04",
    "Mai": "05", "Jun": "06", "Jul": "07", "Ago": "08",
    "Set": "09", "Out": "10", "Nov": "11", "Dez": "12",
}

MESES_EN = {
    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
}


def converter_data_excel_para_stur(data_excel: str) -> str:
    """Converte datas do Excel para dd/mm/aaaa quando possível."""
    texto = str(data_excel or "").strip()
    if not texto or texto.lower() == "nan":
        return ""

    # Ex: 28/Mar/2026
    partes = texto.split("/")
    if len(partes) == 3:
        dia, mes, ano = partes
        mes_num = MESES_PT.get(mes.strip().title(), mes.strip())
        return f"{dia.strip().zfill(2)}/{mes_num.zfill(2)}/{ano.strip()}"

    # Ex: pandas Timestamp ou string ISO
    try:
        dt = pd.to_datetime(texto, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%d/%m/%Y")
    except Exception:
        pass

    return texto


class ExcelService:
    def __init__(self, config: AppConfig):
        self.config = config

    def carregar_transacoes(self, arquivo: Path) -> tuple[pd.DataFrame, str]:
        if not arquivo.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")

        extensao = arquivo.suffix.lower()

        if extensao == ".csv":
            df = self._ler_csv_com_encoding(arquivo)
            return self._normalizar_df(df), "CSV"

        if extensao in {".xlsx", ".xls"}:
            sheet_name = self._definir_aba_transacoes(arquivo)
            header_row = self._encontrar_linha_cabecalho(arquivo, sheet_name)
            df = pd.read_excel(arquivo, sheet_name=sheet_name, dtype=str, header=header_row)
            return self._normalizar_df(df), str(sheet_name)

        raise ValueError(f"Formato não suportado: {extensao}")

    def montar_transacoes(self, df: pd.DataFrame, origem_arquivo: str | None = None) -> list[Transacao]:
        tipo_layout = self._identificar_layout(df)

        coluna_estabelecimento = self._resolver_coluna_estabelecimento(df, tipo_layout=tipo_layout)
        coluna_data = self._resolver_coluna_data_aprovacao(df, tipo_layout=tipo_layout)
        coluna_valor = self._resolver_coluna_valor(df, tipo_layout=tipo_layout)
        coluna_vcn = self._resolver_coluna_vcn(df)
        coluna_extrato = self._resolver_coluna_extrato(df)
        coluna_autorizacao = self._resolver_coluna_autorizacao(df)

        transacoes: list[Transacao] = []

        coluna_obs = self._resolver_coluna_observacao(df)
        coluna_res = self.config.coluna_resultado
        for index, row in df.iterrows():
            estabelecimento = str(row.get(coluna_estabelecimento, "") or "").strip()
            if not estabelecimento or estabelecimento.lower() == "nan":
                continue

            # Linhas com OBSERVAÇÃO preenchida são de Hotelaria — processadas separadamente
            if coluna_obs:
                obs = str(row.get(coluna_obs, "") or "").strip()
                if obs and obs.lower() != "nan":
                    continue

            # Decide se a linha deve ser pulada ou incluída com flag de venda já feita
            venda_ja_ok = False
            resultado_venda_anterior = ""
            if coluna_res in df.columns:
                resultado_existente = str(row.get(coluna_res, "") or "").strip()
                if resultado_existente and resultado_existente.lower() not in ("", "nan"):
                    r = resultado_existente.upper()
                    if "FATURADO" in r:
                        continue
                    if r.startswith("OK") and "OK CONFERÊNCIA" in r:
                        continue  # completamente processado — pula
                    if r.startswith("OK VENDAS"):
                        # Vendas OK mas conferência falhou — inclui só para fase 2
                        venda_ja_ok = True
                        # Preserva tudo antes do primeiro "ERRO" para reescrever limpo depois
                        idx_erro = r.find("| ERRO")
                        resultado_venda_anterior = resultado_existente[:idx_erro].strip().rstrip("|").strip() if idx_erro != -1 else resultado_existente
                    elif r.startswith("OK"):
                        continue  # outro OK desconhecido — pula por segurança
                    # ERRO: inclui normalmente para retentar as duas fases

            data_excel = str(row.get(coluna_data, "") or "").strip() if coluna_data else ""
            data_stur = converter_data_excel_para_stur(data_excel)
            valor_excel = self._parse_decimal(row.get(coluna_valor)) if coluna_valor else None
            vcn = str(row.get(coluna_vcn, "") or "").strip() if coluna_vcn else ""
            extrato_conta = str(row.get(coluna_extrato, "") or "").strip() if coluna_extrato else ""
            data_fatura = self._converter_extrato_para_data_fatura(extrato_conta)

            codigo_venda_vcn = self._extrair_codigo_venda_vcn(vcn)
            companhia = self._identificar_companhia(estabelecimento)
            if companhia == "AZUL":
                localizador_extraido = self._extrair_localizador_azul(estabelecimento)
            else:
                localizador_extraido = self._extrair_localizador_apos_asterisco(estabelecimento)
            termo_busca, coluna_busca, tipo_busca = self._definir_estrategia_busca(
                estabelecimento=estabelecimento,
                codigo_venda_vcn=codigo_venda_vcn,
                localizador_extraido=localizador_extraido,
            )
            codigo_autorizacao = str(row.get(coluna_autorizacao, "") or "").strip() if coluna_autorizacao else ""
            if codigo_autorizacao.lower() == "nan":
                codigo_autorizacao = ""

            transacoes.append(
                Transacao(
                    indice_planilha=index,
                    linha_excel=index + 2,
                    estabelecimento=estabelecimento,
                    data_aprovacao=data_excel,
                    data_stur=data_stur,
                    valor_excel=valor_excel,
                    vcn=vcn,
                    codigo_venda_vcn=codigo_venda_vcn,
                    localizador_extraido=localizador_extraido,
                    termo_busca=termo_busca,
                    coluna_busca=coluna_busca,
                    tipo_busca=tipo_busca,
                    origem_arquivo=origem_arquivo or "",
                    tipo_layout=tipo_layout,
                    extrato_conta=extrato_conta,
                    data_fatura=data_fatura,
                    codigo_autorizacao=codigo_autorizacao,
                    venda_ja_ok=venda_ja_ok,
                    resultado_venda_anterior=resultado_venda_anterior,
                )
            )

        return transacoes

    def escrever_resultado(self, df: pd.DataFrame, transacao: Transacao, resultado: str) -> None:
        coluna = self.config.coluna_resultado
        if coluna not in df.columns:
            df[coluna] = ""
        df.at[transacao.indice_planilha, coluna] = resultado

    def acrescentar_resultado(self, df: pd.DataFrame, transacao: Transacao, resultado: str) -> None:
        coluna = self.config.coluna_resultado
        if coluna not in df.columns:
            df[coluna] = ""
        atual = str(df.at[transacao.indice_planilha, coluna] or "").strip()
        if atual and atual.lower() != "nan":
            df.at[transacao.indice_planilha, coluna] = f"{atual} | {resultado}"
        else:
            df.at[transacao.indice_planilha, coluna] = resultado

    def salvar_saida(self, df: pd.DataFrame, arquivo_original: Path) -> Path:
        self.config.output_dir.mkdir(exist_ok=True)
        saida = self.config.output_dir / f"{arquivo_original.stem}_processado.xlsx"
        df.to_excel(saida, index=False)
        return saida

    def salvar_no_local_com_cores(self, df: pd.DataFrame, arquivo_original: Path) -> Path:
        """
        Salva o DataFrame com cores de linha de volta na pasta de origem:
          - verde  → linha com resultado OK
          - vermelho → linha com resultado ERRO
          - laranja → linha JÁ FATURADO

        Se o arquivo original for CSV, salva como .xlsx de mesmo nome na mesma pasta.
        """
        pasta = arquivo_original.parent
        pasta.mkdir(parents=True, exist_ok=True)

        if arquivo_original.suffix.lower() == ".csv":
            saida = pasta / (arquivo_original.stem + ".xlsx")
        else:
            saida = arquivo_original

        df.to_excel(saida, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(saida)
        ws = wb.active

        header = [str(cell.value or "").strip() for cell in ws[1]]
        coluna = self.config.coluna_resultado
        try:
            col_idx = header.index(coluna) + 1  # 1-based
        except ValueError:
            wb.save(saida)
            return saida

        for row_idx in range(2, ws.max_row + 1):
            resultado = str(ws.cell(row=row_idx, column=col_idx).value or "").strip().upper()
            if "ERRO" in resultado:
                fill = _RED_FILL
            elif resultado.startswith("OK"):
                fill = _GREEN_FILL
            elif "FATURADO" in resultado:
                fill = _ORANGE_FILL
            else:
                continue

            for c in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=c).fill = fill

        wb.save(saida)
        return saida


    def montar_transacoes_hoteis(self, df: pd.DataFrame, origem_arquivo: str | None = None) -> list[TransacaoHotel]:
        """Retorna transações de Hotelaria: linhas onde col V (OBSERVAÇÃO) está preenchida."""
        from datetime import datetime

        data_fatura_hoje = datetime.today().strftime("%d/%m/%Y")

        coluna_obs = self._resolver_coluna_observacao(df)
        if not coluna_obs:
            return []

        tipo_layout = self._identificar_layout(df)
        coluna_estabelecimento = self._resolver_coluna_estabelecimento(df, tipo_layout=tipo_layout)
        coluna_data = self._resolver_coluna_data_aprovacao(df, tipo_layout=tipo_layout)
        coluna_valor = self._resolver_coluna_valor(df, tipo_layout=tipo_layout)
        coluna_autorizacao = self._resolver_coluna_autorizacao(df)
        coluna_titular = self._resolver_coluna_titular(df)
        coluna_cliente = self._resolver_coluna_cliente(df)
        coluna_res = self.config.coluna_resultado

        transacoes: list[TransacaoHotel] = []

        for index, row in df.iterrows():
            observacao = str(row.get(coluna_obs, "") or "").strip()
            if not observacao or observacao.lower() == "nan":
                continue

            if coluna_res in df.columns:
                resultado_existente = str(row.get(coluna_res, "") or "").strip()
                if resultado_existente and resultado_existente.lower() not in ("", "nan"):
                    r = resultado_existente.upper()
                    if r.startswith("OK") or "FATURADO" in r:
                        continue

            estabelecimento = str(row.get(coluna_estabelecimento, "") or "").strip()
            data_excel = str(row.get(coluna_data, "") or "").strip() if coluna_data else ""
            valor_excel = self._parse_decimal(row.get(coluna_valor)) if coluna_valor else None

            codigo_autorizacao = str(row.get(coluna_autorizacao, "") or "").strip() if coluna_autorizacao else ""
            if codigo_autorizacao.lower() == "nan":
                codigo_autorizacao = ""

            titular = str(row.get(coluna_titular, "") or "").strip() if coluna_titular else ""
            if titular.lower() == "nan":
                titular = ""

            cliente = str(row.get(coluna_cliente, "") or "").strip() if coluna_cliente else ""
            if cliente.lower() == "nan":
                cliente = ""

            transacoes.append(
                TransacaoHotel(
                    indice_planilha=index,
                    linha_excel=index + 2,
                    estabelecimento=estabelecimento,
                    data_aprovacao=data_excel,
                    valor_excel=valor_excel,
                    codigo_autorizacao=codigo_autorizacao,
                    titular=titular,
                    observacao=observacao,
                    cliente=cliente,
                    data_fatura=data_fatura_hoje,
                    origem_arquivo=origem_arquivo or "",
                )
            )

        return transacoes

    def escrever_discrepancia_hotel(
        self,
        df: pd.DataFrame,
        transacao: TransacaoHotel,
        valor_tabela: "Decimal | None",
        diferenca: "Decimal | None",
    ) -> None:
        """Escreve valor_tabela (col X) e diferença (col Y) para linhas de hotel com discrepância."""
        cols = list(df.columns)
        if len(cols) > 23:
            df.at[transacao.indice_planilha, cols[23]] = float(valor_tabela) if valor_tabela is not None else ""
        if len(cols) > 24:
            df.at[transacao.indice_planilha, cols[24]] = float(diferenca) if diferenca is not None else ""

    def obter_vencimento_capa(self, arquivo: Path) -> str | None:
        """
        Busca na aba "Capa" o campo "Vencimento" e retorna a data no formato dd/mm/aaaa.

        Layout esperado na fatura:
        Vencimento    22/04/2026
        """
        import logging
        log = logging.getLogger("robo_stur")

        if arquivo.suffix.lower() not in {".xlsx", ".xls"}:
            return None

        excel = pd.ExcelFile(arquivo)
        log.info("[Capa] Abas encontradas: %s", excel.sheet_names)
        sheet_name = None

        for aba in excel.sheet_names:
            if self._normalizar_texto(aba) == "capa":
                sheet_name = aba
                break

        if sheet_name is None:
            log.info("[Capa] Aba 'Capa' não encontrada; usando aba 0 (%s).", excel.sheet_names[0] if excel.sheet_names else "?")
            sheet_name = 0

        df_capa = pd.read_excel(arquivo, sheet_name=sheet_name, header=None, dtype=object)
        log.info("[Capa] Lendo aba '%s' — shape %s", sheet_name, df_capa.shape)

        for row_idx in range(df_capa.shape[0]):
            for col_idx in range(df_capa.shape[1]):
                valor = df_capa.iat[row_idx, col_idx]
                normalizado = self._normalizar_texto(valor)
                if "vencimento" in normalizado:
                    log.info("[Capa] Célula com 'vencimento' em [%d,%d]: '%s'", row_idx, col_idx, valor)
                    # Primeiro tenta a célula ao lado.
                    if col_idx + 1 < df_capa.shape[1]:
                        prox = df_capa.iat[row_idx, col_idx + 1]
                        log.info("[Capa] Célula adjacente [%d,%d]: '%s'", row_idx, col_idx + 1, prox)
                        vencimento = self._formatar_data_saida(prox)
                        if vencimento:
                            log.info("[Capa] Vencimento encontrado: %s", vencimento)
                            return vencimento

                    # Fallback: procura qualquer data na mesma linha.
                    for prox_col_idx in range(col_idx + 1, df_capa.shape[1]):
                        prox = df_capa.iat[row_idx, prox_col_idx]
                        vencimento = self._formatar_data_saida(prox)
                        if vencimento:
                            log.info("[Capa] Vencimento encontrado (fallback col %d): %s", prox_col_idx, vencimento)
                            return vencimento

        log.warning("[Capa] Campo 'Vencimento' não encontrado na aba '%s'.", sheet_name)
        return None

    def _formatar_data_saida(self, valor) -> str | None:
        if valor is None:
            return None

        texto = str(valor).strip()
        if not texto or texto.lower() == "nan":
            return None

        data = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.isna(data):
            return None

        return data.strftime("%d/%m/%Y")

    def validar_total_primeira_aba(self, arquivo: Path, df_processado: pd.DataFrame) -> str:
        if arquivo.suffix.lower() not in {".xlsx", ".xls"}:
            return "Validação de total não aplicada para CSV."

        primeira_aba = pd.read_excel(arquivo, sheet_name=0, header=None, dtype=str)
        total_a_pagar = None

        for _, row in primeira_aba.iterrows():
            valores = [str(v).strip() for v in row.tolist() if str(v).strip() and str(v) != "nan"]
            linha_texto = " ".join(valores).lower()
            if "total a pagar" in linha_texto:
                total_a_pagar = self._primeiro_decimal_na_linha(valores)
                break

        if total_a_pagar is None:
            return "Não encontrei 'Total a pagar' na primeira aba."

        return f"Total a pagar localizado na primeira aba: {total_a_pagar}"


    def _ler_csv_com_encoding(self, arquivo: Path) -> pd.DataFrame:
        """
        Lê CSV tentando os encodings mais comuns.

        O arquivo novo da Clara veio em latin1/cp1252 e separado por ponto e vírgula.
        Mantemos sep=None para também aceitar CSV separado por vírgula.
        """
        ultimo_erro: Exception | None = None

        # Tenta primeiro ponto e vírgula porque é o layout recebido da Clara.
        # Depois tenta autodetect para manter compatibilidade com outros CSVs.
        tentativas = [
            {"sep": ";", "engine": "python"},
            {"sep": None, "engine": "python"},
            {"sep": ",", "engine": "python"},
        ]

        for encoding in ("utf-8-sig", "utf-8", "latin1", "cp1252"):
            for tentativa in tentativas:
                try:
                    df = pd.read_csv(arquivo, dtype=str, encoding=encoding, **tentativa)

                    # Se veio tudo em uma coluna contendo ';', provavelmente o separador foi lido errado.
                    if len(df.columns) == 1 and ";" in str(df.columns[0]):
                        continue

                    return df
                except UnicodeDecodeError as exc:
                    ultimo_erro = exc
                    break
                except Exception as exc:
                    ultimo_erro = exc
                    continue

        raise ValueError(f"Não consegui ler o CSV {arquivo}. Último erro: {ultimo_erro}")

    def _identificar_layout(self, df: pd.DataFrame) -> str:
        """
        Identifica o tipo de arquivo de entrada.

        - CLARA: CSV novo com colunas Transação, Valor original e Valor em R$.
        - PADRAO: planilha/arquivo antigo que já vínhamos usando.
        """
        colunas_normalizadas = {self._normalizar_texto(col) for col in df.columns}

        if "transacao" in colunas_normalizadas and (
            "valor original" in colunas_normalizadas or "valor em r$" in colunas_normalizadas
        ):
            return "CLARA"

        return "PADRAO"

    def _definir_estrategia_busca(
        self,
        estabelecimento: str,
        codigo_venda_vcn: str | None,
        localizador_extraido: str | None,
    ) -> tuple[str, str, str]:
        """
        Define a primeira estratégia da linha.

        Regras:
        - Se o VCN trouxer venda, busca direto por Venda.
        - Se for LATAM, GOL ou AZUL e tiver localizador extraído, busca direto
          pelo Localizador.
        - Se NÃO for nenhuma dessas, mas tiver '*', mantém o estabelecimento como
          termo principal, porém sinaliza que também deve tentar o Localizador extraído.
        - Caso contrário, fluxo genérico normal.
        """
        if codigo_venda_vcn:
            return codigo_venda_vcn, "Venda", "VCN"

        companhia = self._identificar_companhia(estabelecimento)
        if localizador_extraido and companhia:
            return localizador_extraido, "Localizador", companhia

        if localizador_extraido:
            return estabelecimento, "Fornecedor", "GENERICO_COM_LOCALIZADOR"

        return estabelecimento, "Fornecedor", "GENERICO"

    def _identificar_companhia(self, estabelecimento: str) -> str | None:
        texto = str(estabelecimento or "").lower()
        if "latam" in texto:
            return "LATAM"
        if "gol" in texto:
            return "GOL"
        if "azul" in texto:
            return "AZUL"
        return None

    def _extrair_localizador_apos_asterisco(self, estabelecimento: str) -> str | None:
        """Usado por LATAM e GOL — o localizador vem logo depois do '*'."""
        texto = str(estabelecimento or "").strip()
        if "*" not in texto:
            return None

        match = re.search(r"\*\s*([A-Za-z0-9]{6})", texto)
        if not match:
            return None

        return match.group(1).upper()

    def _extrair_localizador_azul(self, estabelecimento: str) -> str | None:
        """Na AZUL o localizador são os últimos 6 caracteres alfanuméricos da descrição."""
        texto = str(estabelecimento or "").strip()
        match = re.search(r"([A-Za-z0-9]{6})\s*$", texto)
        if not match:
            return None

        return match.group(1).upper()

    def _extrair_codigo_venda_vcn(self, vcn: str) -> str | None:
        """
        Cenário futuro explicado pelo cliente:
        VCN pode vir com algo como 'Venda 2238' ou texto contendo o número da venda.
        """
        texto = str(vcn or "").strip()
        if not texto or texto.lower() == "nan":
            return None

        match = re.search(r"venda\D*(\d{3,})", texto, flags=re.IGNORECASE)
        if match:
            return match.group(1)

        # fallback conservador: se o VCN for só número, assume venda
        if re.fullmatch(r"\d{3,}", texto):
            return texto

        return None

    def _encontrar_linha_cabecalho(self, arquivo: Path, sheet_name: str | int) -> int:
        amostra = pd.read_excel(arquivo, sheet_name=sheet_name, header=None, nrows=12, dtype=str)
        palavras_chave = {"estabelecimento", "valor em r$", "data de aprovação", "vcn"}
        for i, row in amostra.iterrows():
            valores = [str(v).strip().lower() for v in row if str(v).strip() and str(v) != "nan"]
            if any(any(kw in v for kw in palavras_chave) for v in valores):
                return i
        return 0

    def _definir_aba_transacoes(self, arquivo: Path) -> str | int:
        if self.config.excel_sheet_transacoes:
            return self.config.excel_sheet_transacoes

        excel = pd.ExcelFile(arquivo)
        for aba in excel.sheet_names:
            if self._normalizar_texto(aba) in {"transacoes", "transacao", "transactions"}:
                return aba
        return 0

    def _normalizar_df(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = [str(col).strip() for col in df.columns]
        df = df.dropna(how="all").reset_index(drop=True)
        return df

    def _resolver_coluna_estabelecimento(self, df: pd.DataFrame, tipo_layout: str = "PADRAO") -> str:
        if tipo_layout == "CLARA":
            candidatos = ["transação", "transacao"]
        else:
            candidatos = ["estabelecimento", "fornecedor", "descricao", "descrição", "historico", "histórico", "transação", "transacao"]

        return self._procurar_coluna(df, candidatos, obrigatoria=True, finalidade="descrição/estabelecimento")

    def _resolver_coluna_data_aprovacao(self, df: pd.DataFrame, tipo_layout: str = "PADRAO") -> str | None:
        if tipo_layout == "CLARA":
            candidatos = ["data da transação", "data da transacao"]
        else:
            candidatos = ["data de aprovação", "data de aprovacao", "data aprovacao", "data aprovação", "data da transação", "data da transacao", "data"]

        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="data")

    def _resolver_coluna_valor(self, df: pd.DataFrame, tipo_layout: str = "PADRAO") -> str | None:
        if tipo_layout == "CLARA":
            # Preferimos Valor em R$ porque já vem convertido. Se não existir, usa Valor original.
            candidatos = ["valor em r$", "valor original"]
        else:
            candidatos = ["valor em r$", "valor em reais", "valor original", "valor", "valor pago", "total", "total a pagar"]

        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="valor")

    def _resolver_coluna_vcn(self, df: pd.DataFrame) -> str | None:
        candidatos = ["vcn", "venda", "codigo venda", "código venda"]
        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="VCN")

    def _resolver_coluna_extrato(self, df: pd.DataFrame) -> str | None:
        candidatos = ["extrato da conta", "extrato", "periodo", "período", "period", "fatura"]
        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="extrato da conta")

    def _resolver_coluna_autorizacao(self, df: pd.DataFrame) -> str | None:
        candidatos = [
            "código de autorização", "codigo de autorizacao",
            "cod. autorização", "cod autorizacao", "cod. autorizacao",
            "código autorização", "codigo autorizacao",
            "autorização", "autorizacao",
        ]
        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="código de autorização")

    def _converter_extrato_para_data_fatura(self, extrato: str) -> str:
        """
        Converte 'Extrato da conta' como '16 May 2026 - 15 Jun 2026' para a data final: '15/06/2026'.
        Essa data é usada como Fatura ao criar conferências LATAM.
        """
        texto = str(extrato or "").strip()
        if not texto or texto.lower() == "nan":
            return ""

        parte_fim = texto.split(" - ")[-1].strip() if " - " in texto else texto

        try:
            dt = pd.to_datetime(parte_fim, dayfirst=True, errors="coerce")
            if pd.notna(dt):
                return dt.strftime("%d/%m/%Y")
        except Exception:
            pass

        match = re.match(r"(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})", parte_fim)
        if match:
            dia, mes_en, ano = match.group(1), match.group(2).title(), match.group(3)
            mes_num = MESES_EN.get(mes_en, "")
            if mes_num:
                return f"{dia.zfill(2)}/{mes_num}/{ano}"

        return ""

    def _procurar_coluna(self, df: pd.DataFrame, candidatos: list[str], obrigatoria: bool, finalidade: str) -> str | None:
        mapa = {self._normalizar_texto(col): col for col in df.columns}
        for candidato in candidatos:
            normalizado = self._normalizar_texto(candidato)
            if normalizado in mapa:
                return mapa[normalizado]

        if obrigatoria:
            raise ValueError(
                f"Não consegui identificar a coluna de {finalidade}. Colunas encontradas: {list(df.columns)}"
            )
        return None

    def _parse_decimal(self, value) -> Decimal | None:
        if value is None:
            return None
        texto = str(value).strip()
        if not texto or texto.lower() == "nan":
            return None
        texto = texto.replace("R$", "").replace(" ", "")
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        texto = re.sub(r"[^0-9.-]", "", texto)
        if not texto:
            return None
        try:
            return Decimal(texto)
        except InvalidOperation:
            return None

    def _primeiro_decimal_na_linha(self, valores: list[str]) -> Decimal | None:
        for valor in valores:
            parsed = self._parse_decimal(valor)
            if parsed is not None:
                return parsed
        return None

    def _resolver_coluna_observacao(self, df: pd.DataFrame) -> str | None:
        candidatos = ["observacao", "observação", "obs", "obs."]
        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="observação")

    def _resolver_coluna_titular(self, df: pd.DataFrame) -> str | None:
        candidatos = ["titular", "titular do cartao", "titular do cartão", "nome"]
        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="titular")

    def _resolver_coluna_cliente(self, df: pd.DataFrame) -> str | None:
        candidatos = ["empresa", "cliente", "company", "nome empresa"]
        return self._procurar_coluna(df, candidatos, obrigatoria=False, finalidade="cliente/empresa")

    def _normalizar_texto(self, texto: str) -> str:
        return (
            str(texto).strip().lower()
            .replace("ç", "c")
            .replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
            .replace("é", "e").replace("ê", "e")
            .replace("í", "i")
            .replace("ó", "o").replace("ô", "o").replace("õ", "o")
            .replace("ú", "u")
        )
